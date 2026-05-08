from __future__ import annotations

import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from cleaning_rules import clean_master_title, clean_title, drop_disabled_rows, extract_master_work_title, text


S2_TITLE_COL_CAND = ["콘텐츠명", "콘텐츠 제목", "Title", "ContentName", "제목"]
S2_ID_COL_CAND = [
    "판매채널콘텐츠ID",
    "판매채널컨텐츠ID",
    "판매채널 콘텐츠ID",
    "판매채널 컨텐츠ID",
    "schnCtnsId",
    "SchnCtnsId",
    "SalesChannelContentID",
    "ContentID",
    "ID",
]
SETTLEMENT_TITLE_COL_CAND = [
    "컨텐츠",
    "타이틀",
    "작품명",
    "도서명",
    "작품 제목",
    "상품명",
    "이용상품명",
    "상품 제목",
    "ProductName",
    "Title",
    "제목",
    "컨텐츠명",
    "콘텐츠명",
    "시리즈명",
]
MASTER_TITLE_COL_CAND = ["콘텐츠명", "콘텐츠 제목", "Title", "ContentName", "제목"]
MASTER_ID_COL_CAND = ["콘텐츠ID", "판매채널콘텐츠ID", "ID", "ContentID"]

MATCH_OK = "matched"
MATCH_NONE = "no_match"
MATCH_AMBIGUOUS = "ambiguous"
MATCH_BLANK = "blank_key"
MATCH_SKIPPED = "skipped"


@dataclass
class MappingResult:
    rows: pd.DataFrame
    summary: pd.DataFrame
    review_rows: pd.DataFrame
    duplicate_candidates: pd.DataFrame
    input_validation: pd.DataFrame


def pick_column(candidates: Iterable[str], df: pd.DataFrame, label: str) -> str:
    for column in candidates:
        if column in df.columns:
            return column
    available = ", ".join(map(str, df.columns))
    expected = ", ".join(candidates)
    raise ValueError(f"{label} 컬럼을 찾지 못했습니다. 기대 컬럼: {expected}. 현재 컬럼: {available}")


def read_first_sheet(source: Any) -> pd.DataFrame:
    return pd.read_excel(source, dtype=object, engine="openpyxl")


def read_all_sheets(source: Any) -> pd.DataFrame:
    sheets = pd.read_excel(source, sheet_name=None, dtype=object, engine="openpyxl")
    frames: list[pd.DataFrame] = []
    for sheet_name, frame in sheets.items():
        copied = frame.copy()
        copied["source_sheet"] = sheet_name
        frames.append(copied)
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def load_master(path: Path) -> pd.DataFrame:
    return read_first_sheet(path)


def _join_unique(values: Iterable[Any], limit: int = 30) -> str:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        cleaned = text(value)
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        result.append(cleaned)
        if len(result) >= limit:
            break
    return " | ".join(result)


def _candidate_index(
    df: pd.DataFrame,
    *,
    source: str,
    key_col: str,
    id_col: str,
    title_col: str,
    extra_cols: list[str] | None = None,
) -> tuple[pd.DataFrame, dict[str, dict[str, str]]]:
    extra_cols = extra_cols or []
    rows: list[dict[str, Any]] = []
    lookup: dict[str, dict[str, str]] = {}

    working = df.copy()
    working[id_col] = working[id_col].map(text)
    working[title_col] = working[title_col].map(text)

    for key, group in working[working[key_col].map(bool)].groupby(key_col, dropna=False):
        ids = [value for value in dict.fromkeys(group[id_col].map(text)) if value]
        titles = [value for value in dict.fromkeys(group[title_col].map(text)) if value]
        status = MATCH_OK if len(ids) == 1 else MATCH_AMBIGUOUS
        row = {
            "source": source,
            "정제키": key,
            "매칭상태": status,
            "후보행수": len(group),
            "후보ID수": len(ids),
            "후보ID목록": " | ".join(ids[:30]),
            "후보콘텐츠명목록": " | ".join(titles[:30]),
        }
        for col in extra_cols:
            if col in group.columns:
                row[f"{col}목록"] = _join_unique(group[col])
        rows.append(row)
        lookup[key] = row

    candidates = pd.DataFrame(rows)
    return candidates, lookup


def _status_for(key: str, index: dict[str, dict[str, str]]) -> str:
    if not key:
        return MATCH_BLANK
    if key not in index:
        return MATCH_NONE
    return text(index[key].get("매칭상태"))


def _value_for(key: str, index: dict[str, dict[str, str]], field: str) -> str:
    if key not in index:
        return ""
    return text(index[key].get(field))


def _single_id_for(key: str, index: dict[str, dict[str, str]]) -> str:
    if _status_for(key, index) != MATCH_OK:
        return ""
    ids = _value_for(key, index, "후보ID목록").split(" | ")
    return text(ids[0]) if ids else ""


def _single_title_for(key: str, index: dict[str, dict[str, str]]) -> str:
    if _status_for(key, index) != MATCH_OK:
        return ""
    titles = _value_for(key, index, "후보콘텐츠명목록").split(" | ")
    return text(titles[0]) if titles else ""


def _single_extra_for(key: str, index: dict[str, dict[str, str]], field: str) -> str:
    if _status_for(key, index) != MATCH_OK:
        return ""
    values = _value_for(key, index, field).split(" | ")
    return text(values[0]) if values else ""


def _review_reason(row: pd.Series) -> str:
    reasons: list[str] = []
    if row["정제_상품명"] == "":
        reasons.append("정산서 상품명 정제키 없음")
    if row["S2_매칭상태"] == MATCH_NONE:
        reasons.append("S2 미매핑")
    elif row["S2_매칭상태"] == MATCH_AMBIGUOUS:
        reasons.append("S2 중복 후보")
    elif row["S2_매칭상태"] == MATCH_BLANK:
        reasons.append("S2 매칭 불가")
    if row["IPS_매칭상태"] == MATCH_NONE:
        reasons.append("IPS 미매핑")
    elif row["IPS_매칭상태"] == MATCH_AMBIGUOUS:
        reasons.append("IPS 중복 후보")
    elif row["IPS_매칭상태"] == MATCH_BLANK:
        reasons.append("IPS 매칭 불가")
    return " | ".join(reasons)


def build_mapping(
    s2_df: pd.DataFrame,
    settlement_df: pd.DataFrame,
    master_df: pd.DataFrame | None = None,
) -> MappingResult:
    s2_df = drop_disabled_rows(s2_df)
    master_df = drop_disabled_rows(master_df) if master_df is not None else None

    s2_title_col = pick_column(S2_TITLE_COL_CAND, s2_df, "S2 콘텐츠명")
    s2_id_col = pick_column(S2_ID_COL_CAND, s2_df, "S2 판매채널콘텐츠ID")
    settlement_title_col = pick_column(SETTLEMENT_TITLE_COL_CAND, settlement_df, "정산서 상품명")
    use_ips = master_df is not None and not master_df.empty
    master_title_col = pick_column(MASTER_TITLE_COL_CAND, master_df, "IPS 콘텐츠명") if use_ips else ""
    master_id_col = pick_column(MASTER_ID_COL_CAND, master_df, "IPS 콘텐츠ID") if use_ips else ""

    s2 = s2_df.copy()
    settlement = settlement_df.copy()
    master = master_df.copy() if use_ips else pd.DataFrame()

    s2["_정제키"] = s2[s2_title_col].map(clean_title)
    settlement["_정제키"] = settlement[settlement_title_col].map(clean_title)
    if use_ips:
        master["_추출작품명"] = master[master_title_col].map(extract_master_work_title)
        master["_정제키"] = master[master_title_col].map(clean_master_title)

    s2_candidates, s2_index = _candidate_index(
        s2,
        source="S2",
        key_col="_정제키",
        id_col=s2_id_col,
        title_col=s2_title_col,
        extra_cols=[col for col in ["판매채널명", "판매채널ID", "콘텐츠ID"] if col in s2.columns],
    )
    if use_ips:
        master_candidates, master_index = _candidate_index(
            master,
            source="IPS",
            key_col="_정제키",
            id_col=master_id_col,
            title_col=master_title_col,
            extra_cols=[col for col in ["작가필명", "서비스유형", "담당부서", "담당자명"] if col in master.columns],
        )
    else:
        master_candidates = pd.DataFrame(columns=["source", "정제키", "매칭상태"])
        master_index = {}

    result = pd.DataFrame(
        {
            "정산서_원본행번호": range(1, len(settlement) + 1),
            "정산서_콘텐츠명": settlement[settlement_title_col].map(text),
            "정제_상품명": settlement["_정제키"],
        }
    )
    result["S2_매칭상태"] = result["정제_상품명"].map(lambda key: _status_for(key, s2_index))
    result["S2_판매채널콘텐츠ID"] = result["정제_상품명"].map(lambda key: _single_id_for(key, s2_index))
    result["S2_콘텐츠ID"] = result["정제_상품명"].map(lambda key: _single_extra_for(key, s2_index, "콘텐츠ID목록"))
    result["S2_콘텐츠명"] = result["정제_상품명"].map(lambda key: _single_title_for(key, s2_index))
    result["S2_후보수"] = result["정제_상품명"].map(lambda key: _value_for(key, s2_index, "후보ID수") or "0")
    result["S2_후보ID목록"] = result["정제_상품명"].map(lambda key: _value_for(key, s2_index, "후보ID목록"))
    result["S2_후보콘텐츠명목록"] = result["정제_상품명"].map(lambda key: _value_for(key, s2_index, "후보콘텐츠명목록"))

    if use_ips:
        result["IPS_매칭상태"] = result["정제_상품명"].map(lambda key: _status_for(key, master_index))
        result["IPS_콘텐츠ID"] = result["정제_상품명"].map(lambda key: _single_id_for(key, master_index))
        result["IPS_콘텐츠명"] = result["정제_상품명"].map(lambda key: _single_title_for(key, master_index))
        result["IPS_후보수"] = result["정제_상품명"].map(lambda key: _value_for(key, master_index, "후보ID수") or "0")
        result["IPS_후보ID목록"] = result["정제_상품명"].map(lambda key: _value_for(key, master_index, "후보ID목록"))
        result["IPS_후보콘텐츠명목록"] = result["정제_상품명"].map(lambda key: _value_for(key, master_index, "후보콘텐츠명목록"))
    else:
        result["IPS_매칭상태"] = MATCH_SKIPPED
        result["IPS_콘텐츠ID"] = ""
        result["IPS_콘텐츠명"] = ""
        result["IPS_후보수"] = "0"
        result["IPS_후보ID목록"] = ""
        result["IPS_후보콘텐츠명목록"] = ""

    result["검토필요사유"] = result.apply(_review_reason, axis=1)
    result["검토필요(Y/N)"] = result["검토필요사유"].map(lambda value: "Y" if text(value) else "N")

    original_cols = [col for col in settlement_df.columns if col not in result.columns]
    if original_cols:
        prefixed = settlement_df[original_cols].copy()
        prefixed.columns = [f"정산서원본_{col}" for col in prefixed.columns]
        result = pd.concat([result, prefixed], axis=1)

    duplicate_candidates = pd.concat(
        [
            s2_candidates[s2_candidates["매칭상태"].eq(MATCH_AMBIGUOUS)],
            master_candidates[master_candidates["매칭상태"].eq(MATCH_AMBIGUOUS)],
        ],
        ignore_index=True,
    )
    review_rows = result[result["검토필요(Y/N)"].eq("Y")].copy()
    input_validation = _build_input_validation(
        s2,
        settlement,
        master,
        s2_title_col=s2_title_col,
        s2_id_col=s2_id_col,
        settlement_title_col=settlement_title_col,
        master_title_col=master_title_col,
        master_id_col=master_id_col,
        s2_candidates=s2_candidates,
        master_candidates=master_candidates,
    )
    summary = _build_summary(result, duplicate_candidates, input_validation)
    return MappingResult(
        rows=result,
        summary=summary,
        review_rows=review_rows,
        duplicate_candidates=duplicate_candidates,
        input_validation=input_validation,
    )


def _build_input_validation(
    s2: pd.DataFrame,
    settlement: pd.DataFrame,
    master: pd.DataFrame,
    *,
    s2_title_col: str,
    s2_id_col: str,
    settlement_title_col: str,
    master_title_col: str,
    master_id_col: str,
    s2_candidates: pd.DataFrame,
    master_candidates: pd.DataFrame,
) -> pd.DataFrame:
    rows = [
        ("S2 행 수", len(s2)),
        ("S2 콘텐츠명 컬럼", s2_title_col),
        ("S2 ID 컬럼", s2_id_col),
        ("S2 빈 정제키 행 수", int(s2["_정제키"].eq("").sum())),
        ("S2 중복 후보 정제키 수", int(s2_candidates["매칭상태"].eq(MATCH_AMBIGUOUS).sum())),
        ("정산서 행 수", len(settlement)),
        ("정산서 상품명 컬럼", settlement_title_col),
        ("정산서 빈 정제키 행 수", int(settlement["_정제키"].eq("").sum())),
    ]
    if master_title_col:
        rows.extend(
            [
                ("IPS 행 수", len(master)),
                ("IPS 콘텐츠명 컬럼", master_title_col),
                ("IPS ID 컬럼", master_id_col),
                ("IPS 빈 정제키 행 수", int(master["_정제키"].eq("").sum())),
                ("IPS 중복 후보 정제키 수", int(master_candidates["매칭상태"].eq(MATCH_AMBIGUOUS).sum())),
            ]
        )
    else:
        rows.append(("IPS 검산", "skipped"))
    if "귀속법인" in master.columns:
        rows.append(("IPS 귀속법인", _join_unique(master["귀속법인"], limit=10)))
    if "콘텐츠형태" in master.columns:
        rows.append(("IPS 콘텐츠형태", _join_unique(master["콘텐츠형태"], limit=10)))
    return pd.DataFrame(rows, columns=["항목", "값"])


def _build_summary(
    result: pd.DataFrame,
    duplicate_candidates: pd.DataFrame,
    input_validation: pd.DataFrame,
) -> pd.DataFrame:
    rows = [
        ("정산서 행 수", len(result)),
        ("검토필요 행 수", int(result["검토필요(Y/N)"].eq("Y").sum())),
        ("S2 matched", int(result["S2_매칭상태"].eq(MATCH_OK).sum())),
        ("S2 ambiguous", int(result["S2_매칭상태"].eq(MATCH_AMBIGUOUS).sum())),
        ("S2 no_match", int(result["S2_매칭상태"].eq(MATCH_NONE).sum())),
        ("S2 콘텐츠ID present", int(result["S2_콘텐츠ID"].map(text).ne("").sum())),
        ("IPS matched", int(result["IPS_매칭상태"].eq(MATCH_OK).sum())),
        ("IPS ambiguous", int(result["IPS_매칭상태"].eq(MATCH_AMBIGUOUS).sum())),
        ("IPS no_match", int(result["IPS_매칭상태"].eq(MATCH_NONE).sum())),
        ("IPS skipped", int(result["IPS_매칭상태"].eq(MATCH_SKIPPED).sum())),
        ("중복 후보 정제키 수", len(duplicate_candidates)),
    ]
    validation_map = dict(zip(input_validation["항목"], input_validation["값"]))
    for key in ("IPS 행 수", "IPS 귀속법인", "IPS 콘텐츠형태"):
        if key in validation_map:
            rows.append((key, validation_map[key]))
    return pd.DataFrame(rows, columns=["항목", "값"])


def export_mapping(result: MappingResult) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        result.summary.to_excel(writer, sheet_name="요약", index=False)
        result.input_validation.to_excel(writer, sheet_name="입력검증", index=False)
        result.rows.to_excel(writer, sheet_name="행별매핑결과", index=False)
        result.review_rows.to_excel(writer, sheet_name="검토필요", index=False)
        result.duplicate_candidates.to_excel(writer, sheet_name="중복후보", index=False)

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        warning_fill = PatternFill("solid", fgColor="FFF2CC")
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            if ws.max_row and ws.max_column:
                ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if ws.title in {"행별매핑결과", "검토필요"}:
                reason_col = None
                for idx, cell in enumerate(ws[1], start=1):
                    if cell.value == "검토필요(Y/N)":
                        reason_col = idx
                        break
                if reason_col:
                    for row in ws.iter_rows(min_row=2):
                        if text(row[reason_col - 1].value) == "Y":
                            for cell in row:
                                cell.fill = warning_fill
            for idx, column_cells in enumerate(ws.columns, start=1):
                max_len = max(len(text(cell.value)) for cell in column_cells)
                ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 70)
    return buffer.getvalue()
