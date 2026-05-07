from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from mapping_core import MATCH_OK, text


S2_TRANSFER_COLUMNS = ["판매채널콘텐츠ID", "판매금액", "정산기준액", "상계금액"]


@dataclass
class S2TransferResult:
    rows: pd.DataFrame
    blocked_rows: pd.DataFrame
    summary: pd.DataFrame
    exportable: bool
    blocking_messages: list[str]


def build_s2_transfer(
    mapping_rows: pd.DataFrame,
    *,
    amount_policy_locked: bool,
    s2_gate: str,
    allow_blank_offset_as_zero: bool = False,
) -> S2TransferResult:
    missing_columns = [
        column
        for column in [
            "S2_매칭상태",
            "S2_판매채널콘텐츠ID",
            "정산서원본_판매금액_후보",
            "정산서원본_정산기준액_후보",
            "정산서원본_상계금액_후보",
        ]
        if column not in mapping_rows.columns
    ]
    if missing_columns:
        message = f"S2 전송자료 필수 입력 컬럼이 없습니다: {', '.join(missing_columns)}"
        return _blocked_result(mapping_rows, message)

    output_rows: list[dict[str, Any]] = []
    blocked_rows: list[dict[str, Any]] = []
    global_reasons = []
    if not amount_policy_locked:
        global_reasons.append(f"금액 정책 미확정: {s2_gate}")

    for idx, row in mapping_rows.reset_index(drop=True).iterrows():
        reasons = list(global_reasons)
        s2_id = text(row.get("S2_판매채널콘텐츠ID"))
        sale_amount = _number_or_none(row.get("정산서원본_판매금액_후보"))
        settlement_amount = _number_or_none(row.get("정산서원본_정산기준액_후보"))
        offset_raw = row.get("정산서원본_상계금액_후보")
        offset_amount = _number_or_none(offset_raw)

        if text(row.get("S2_매칭상태")) != MATCH_OK:
            reasons.append(f"S2 매칭상태가 matched가 아닙니다: {text(row.get('S2_매칭상태'))}")
        if not s2_id:
            reasons.append("판매채널콘텐츠ID가 비었습니다.")
        if sale_amount is None:
            reasons.append("판매금액 후보가 비었거나 숫자가 아닙니다.")
        if settlement_amount is None:
            reasons.append("정산기준액 후보가 비었거나 숫자가 아닙니다.")
        if offset_amount is None:
            if allow_blank_offset_as_zero and not text(offset_raw):
                offset_amount = 0
            else:
                reasons.append("상계금액 후보가 비었거나 숫자가 아닙니다.")

        if reasons:
            blocked_rows.append(
                {
                    "row_number": idx + 1,
                    "정산서_콘텐츠명": text(row.get("정산서_콘텐츠명")),
                    "S2_판매채널콘텐츠ID": s2_id,
                    "S2_매칭상태": text(row.get("S2_매칭상태")),
                    "판매금액_후보": row.get("정산서원본_판매금액_후보"),
                    "정산기준액_후보": row.get("정산서원본_정산기준액_후보"),
                    "상계금액_후보": offset_raw,
                    "차단사유": " | ".join(reasons),
                }
            )
            continue

        output_rows.append(
            {
                "판매채널콘텐츠ID": s2_id,
                "판매금액": _normalize_number(sale_amount),
                "정산기준액": _normalize_number(settlement_amount),
                "상계금액": _normalize_number(offset_amount),
            }
        )

    rows = pd.DataFrame(output_rows, columns=S2_TRANSFER_COLUMNS)
    blocked = pd.DataFrame(blocked_rows)
    blocking_messages = _blocking_messages(
        amount_policy_locked=amount_policy_locked,
        s2_gate=s2_gate,
        missing_columns=[],
        blocked_rows=len(blocked),
    )
    exportable = amount_policy_locked and not rows.empty and blocked.empty
    summary = pd.DataFrame(
        [
            ("전송 가능", "Y" if exportable else "N"),
            ("전송 후보 행 수", len(rows)),
            ("차단 행 수", len(blocked)),
            ("금액 정책 잠금", "Y" if amount_policy_locked else "N"),
            ("S2 gate", s2_gate),
        ],
        columns=["항목", "값"],
    )
    return S2TransferResult(rows=rows, blocked_rows=blocked, summary=summary, exportable=exportable, blocking_messages=blocking_messages)


def export_s2_transfer(result: S2TransferResult) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        result.summary.to_excel(writer, sheet_name="요약", index=False)
        result.rows.to_excel(writer, sheet_name="S2전송자료", index=False)
        result.blocked_rows.to_excel(writer, sheet_name="차단행", index=False)

        header_fill = PatternFill("solid", fgColor="7030A0")
        header_font = Font(color="FFFFFF", bold=True)
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            if ws.max_row and ws.max_column:
                ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for idx, column_cells in enumerate(ws.columns, start=1):
                max_len = max(len(text(cell.value)) for cell in column_cells)
                ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 70)
    return buffer.getvalue()


def _blocked_result(mapping_rows: pd.DataFrame, message: str) -> S2TransferResult:
    blocked = pd.DataFrame(
        [
            {
                "row_number": idx + 1,
                "정산서_콘텐츠명": text(row.get("정산서_콘텐츠명")),
                "차단사유": message,
            }
            for idx, row in mapping_rows.reset_index(drop=True).iterrows()
        ]
    )
    summary = pd.DataFrame(
        [
            ("전송 가능", "N"),
            ("전송 후보 행 수", 0),
            ("차단 행 수", len(blocked)),
            ("차단사유", message),
        ],
        columns=["항목", "값"],
    )
    return S2TransferResult(
        rows=pd.DataFrame(columns=S2_TRANSFER_COLUMNS),
        blocked_rows=blocked,
        summary=summary,
        exportable=False,
        blocking_messages=[message],
    )


def _blocking_messages(
    *,
    amount_policy_locked: bool,
    s2_gate: str,
    missing_columns: list[str],
    blocked_rows: int,
) -> list[str]:
    messages = []
    if missing_columns:
        messages.append(f"필수 컬럼 누락: {', '.join(missing_columns)}")
    if not amount_policy_locked:
        messages.append(f"S2 전송자료 출력 차단: {s2_gate}")
    if blocked_rows:
        messages.append(f"S2 전송자료 차단 행 {blocked_rows:,}개가 있습니다.")
    return messages


def _number_or_none(value: Any) -> float | None:
    value_text = text(value)
    if not value_text:
        return None
    cleaned = value_text.replace(",", "").replace("원", "").replace(" ", "")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _normalize_number(value: float | None) -> int | float:
    if value is None:
        return 0
    return int(value) if float(value).is_integer() else value
