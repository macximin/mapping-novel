from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from settlement_adapters import REGISTRY, detect_platform, normalize_settlement


DEFAULT_SOURCE_ROOT = Path(r"\\172.16.10.120\소설사업부\판무팀_ssot\100_계산서_매출등록_자료")


def main() -> None:
    args = parse_args()
    root = Path(args.root)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    adapter = adapter_amounts(root, args.period)
    legacy = legacy_amounts(root, args.period)
    rows = []
    for platform in sorted(REGISTRY):
        spec = REGISTRY[platform]
        a = adapter.get(platform, empty_amounts())
        l = legacy.get(platform, empty_amounts())
        row = {
            "platform": platform,
            "period": args.period,
            "adapter_file_count": a["file_count"],
            "legacy_mapping_file_count": l["file_count"],
            "adapter_default_rows": a["rows"],
            "legacy_mapping_rows": l["rows"],
            "adapter_sale_sum": a["sale"],
            "legacy_sale_sum": l["sale"],
            "sale_delta": round(a["sale"] - l["sale"], 6),
            "adapter_settlement_sum": a["settlement"],
            "legacy_settlement_sum": l["settlement"],
            "settlement_delta": round(a["settlement"] - l["settlement"], 6),
            "adapter_offset_sum": a["offset"],
            "legacy_offset_sum": l["offset"],
            "offset_delta": round(a["offset"] - l["offset"], 6),
            "amount_rule_status": spec.amount_rule_status,
            "s2_amount_policy_locked": spec.s2_amount_policy_locked,
            "reconcile_status": reconcile_status(a, l),
            "legacy_files": " | ".join(l["files"][:5]),
        }
        rows.append(row)

    out = out_dir / (args.output or f"{args.period.replace('-', '_')}_legacy_mapping_amount_reconcile.csv")
    write_csv(out, rows)
    print(out)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare adapter amount candidates with legacy mapping workbooks.")
    parser.add_argument("--root", default=str(DEFAULT_SOURCE_ROOT))
    parser.add_argument("--period", required=True)
    parser.add_argument("--out-dir", default="doc/2026-05-07")
    parser.add_argument("--output", default="")
    return parser.parse_args()


def adapter_amounts(root: Path, period: str) -> dict[str, dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = defaultdict(empty_amounts)
    for path in discover_files(root, period, keyword="정산상세"):
        rel = str(path.relative_to(root))
        platform = detect_platform(rel) or Path(rel).parts[0]
        if platform not in REGISTRY:
            continue
        result = normalize_settlement(path, platform=platform, source_name=rel)
        feed = result.default_feed_rows
        bucket = grouped[platform]
        bucket["file_count"] += 1
        bucket["files"].append(rel)
        bucket["rows"] += len(feed)
        bucket["sale"] += numeric_sum(feed.get("판매금액_후보", pd.Series(dtype=object)))
        bucket["settlement"] += numeric_sum(feed.get("정산기준액_후보", pd.Series(dtype=object)))
        bucket["offset"] += numeric_sum(feed.get("상계금액_후보", pd.Series(dtype=object)))
    return dict(grouped)


def legacy_amounts(root: Path, period: str) -> dict[str, dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = defaultdict(empty_amounts)
    for path in discover_files(root, period, keyword="_매핑"):
        rel = str(path.relative_to(root))
        platform = detect_platform(rel) or Path(rel).parts[0]
        if platform not in REGISTRY:
            continue
        frame = read_legacy_sheet(path)
        bucket = grouped[platform]
        bucket["file_count"] += 1
        bucket["files"].append(rel)
        bucket["rows"] += len(frame)
        bucket["sale"] += numeric_sum(pick_col(frame, ["판매금액", "총판매금액", "총금액"]))
        bucket["settlement"] += numeric_sum(pick_col(frame, ["정산금액", "정산 금액"]))
        bucket["offset"] += numeric_sum(pick_col(frame, ["상계금액", "상계"]))
    return dict(grouped)


def discover_files(root: Path, period: str, keyword: str) -> list[Path]:
    result: list[Path] = []
    for path in root.rglob("*.xlsx"):
        if path.name.startswith("~$") or keyword not in path.name:
            continue
        if path_period(path, root) == period:
            result.append(path)
    return sorted(result, key=lambda p: str(p))


def path_period(path: Path, root: Path) -> str:
    rel = str(path.relative_to(root)).replace(" ", "")
    match = re.search(r"(20\d{2})년(\d{1,2})월", rel)
    if match:
        return f"{match.group(1)}-{int(match.group(2)):02d}"
    match = re.search(r"\\(\d{1,2})월\\", str(path.relative_to(root)))
    if match:
        year = "2025" if "25년" in str(path.relative_to(root)) else "2026"
        return f"{year}-{int(match.group(1)):02d}"
    return ""


def read_legacy_sheet(path: Path) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    headers = [str(value).strip() if value is not None else f"column_{idx + 1}" for idx, value in enumerate(rows[0])]
    frame = pd.DataFrame(rows[1:], columns=make_unique(headers))
    id_col = pick_col(frame, ["매핑_판매채널콘텐츠ID", "S2_판매채널콘텐츠ID"])
    if id_col.empty:
        return frame
    return frame[id_col.map(lambda value: str(value).strip() not in {"", "None", "nan"})].copy()


def pick_col(frame: pd.DataFrame, candidates: list[str]) -> pd.Series:
    normalized = {normalize(candidate) for candidate in candidates}
    for column in frame.columns:
        if normalize(str(column)) in normalized:
            return frame[column]
    return pd.Series(dtype=object)


def numeric_sum(series: pd.Series) -> float:
    if series.empty:
        return 0.0
    cleaned = series.astype(str).str.replace(",", "", regex=False).str.replace("원", "", regex=False)
    return float(pd.to_numeric(cleaned, errors="coerce").fillna(0).sum())


def reconcile_status(adapter: dict[str, Any], legacy: dict[str, Any]) -> str:
    if legacy["file_count"] == 0:
        return "no_legacy_mapping"
    if adapter["file_count"] == 0:
        return "no_adapter_source"
    sale_match = abs(adapter["sale"] - legacy["sale"]) < 1
    settlement_match = abs(adapter["settlement"] - legacy["settlement"]) < 1
    offset_match = abs(adapter["offset"] - legacy["offset"]) < 1
    if sale_match and settlement_match and offset_match:
        return "sale_settlement_offset_match"
    if sale_match and settlement_match:
        return "sale_settlement_match"
    if settlement_match:
        return "settlement_match_only"
    return "mismatch_needs_policy_review"


def empty_amounts() -> dict[str, Any]:
    return {"file_count": 0, "rows": 0, "sale": 0.0, "settlement": 0.0, "offset": 0.0, "files": []}


def normalize(value: str) -> str:
    return re.sub(r"[\s_()\[\]]+", "", value).lower()


def make_unique(headers: list[str]) -> list[str]:
    result = []
    seen: dict[str, int] = {}
    for header in headers:
        count = seen.get(header, 0) + 1
        seen[header] = count
        result.append(header if count == 1 else f"{header}__{count}")
    return result


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fields = list(rows[0]) if rows else []
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


if __name__ == "__main__":
    main()
