from __future__ import annotations

import fnmatch
import io
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, BinaryIO, Iterable
from xml.etree import ElementTree as ET

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from cleaning_rules import clean_title, text


OOXML_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OOXML_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
OOXML_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
OOXML_NS = {
    "x": OOXML_MAIN_NS,
    "r": OOXML_REL_NS,
    "pr": OOXML_PACKAGE_REL_NS,
}


STANDARD_TITLE_COLUMN = "상품명"
STANDARD_COLUMNS = [
    "platform",
    "source_file",
    "source_sheet",
    "source_row",
    "file_status",
    "row_status",
    STANDARD_TITLE_COLUMN,
    "작가명",
    "외부콘텐츠ID",
    "판매금액_후보",
    "정산기준액_후보",
    "상계금액_후보",
    "정제_상품명",
    "parser_contract",
    "amount_rule_status",
    "s2_gate",
]

_PLATFORM_BLOCKED_RAW_COLUMN_KEYWORDS = {
    "로망띠끄": ("isbn",),
}


@dataclass(frozen=True)
class AdapterSpec:
    platform: str
    parser_contract: str
    final_class: str
    source_sheet_rule: str
    exclude_rule: str
    title_from: str
    author_from: str
    external_id_from: str
    sale_amount_from: str
    settlement_amount_from: str
    offset_amount_from: str
    amount_rule_status: str
    s2_gate: str

    @property
    def blocks_default_feed(self) -> bool:
        return self.final_class == "non_s2_source_blocked"

    @property
    def s2_amount_policy_locked(self) -> bool:
        return self.amount_rule_status == "candidate_confirmed_after_reconcile"

    @property
    def title_candidates(self) -> list[str]:
        return _field_candidates(self.title_from)

    @property
    def author_candidates(self) -> list[str]:
        return _field_candidates(self.author_from)

    @property
    def external_id_candidates(self) -> list[str]:
        return _field_candidates(self.external_id_from)

    @property
    def sale_amount_candidates(self) -> list[str]:
        return _field_candidates(self.sale_amount_from)

    @property
    def settlement_amount_candidates(self) -> list[str]:
        return _field_candidates(self.settlement_amount_from)

    @property
    def offset_amount_candidates(self) -> list[str]:
        return _field_candidates(self.offset_amount_from)


@dataclass
class SheetAudit:
    sheet: str
    status: str
    header_row: str = ""
    data_start: str = ""
    parsed_rows: int = 0
    title_present_rows: int = 0
    note: str = ""


@dataclass
class NormalizationResult:
    platform: str
    source_name: str
    file_status: str
    spec: AdapterSpec
    rows: pd.DataFrame
    sheet_audits: list[SheetAudit]

    @property
    def default_feed_rows(self) -> pd.DataFrame:
        if self.file_status != "include" or self.spec.blocks_default_feed:
            return self.rows.iloc[0:0].copy()
        if "row_status" not in self.rows.columns:
            return self.rows.iloc[0:0].copy()
        return self.rows[self.rows["row_status"].eq("data")].copy()

    def to_mapping_feed(self) -> pd.DataFrame:
        feed = self.default_feed_rows
        if feed.empty:
            return pd.DataFrame(columns=[STANDARD_TITLE_COLUMN])
        cols = [STANDARD_TITLE_COLUMN, "작가명", "외부콘텐츠ID", "판매금액_후보", "정산기준액_후보", "상계금액_후보"]
        return feed[[col for col in cols if col in feed.columns]].copy()


@dataclass(frozen=True)
class _FallbackMergedRange:
    bounds: tuple[int, int, int, int]


@dataclass
class _FallbackMergedCells:
    ranges: list[_FallbackMergedRange]


@dataclass
class _FallbackWorksheet:
    title: str
    values: list[list[Any]]
    merged_cells: _FallbackMergedCells

    def iter_rows(self, *, values_only: bool = False):
        if not values_only:
            raise ValueError("OOXML fallback worksheets only support values_only=True")
        for row in self.values:
            yield tuple(row)


@dataclass
class _FallbackWorkbook:
    worksheets: list[_FallbackWorksheet]


def adapter_audit_dataframe(result: NormalizationResult) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "sheet": audit.sheet,
                "status": audit.status,
                "header_row": audit.header_row,
                "data_start": audit.data_start,
                "parsed_rows": audit.parsed_rows,
                "title_present_rows": audit.title_present_rows,
                "note": audit.note,
            }
            for audit in result.sheet_audits
        ]
    )


def adapter_blocking_messages(result: NormalizationResult) -> list[str]:
    messages: list[str] = []
    if result.file_status == "excluded_by_rule":
        messages.append(f"파일 제외 규칙에 걸렸습니다: {result.spec.exclude_rule}")
    if result.file_status == "review_gate_not_default":
        messages.append("통합본/확장체크 등 검토용 변형 파일이라 기본 S2 매핑 입력에서 제외했습니다.")
    if result.spec.blocks_default_feed:
        messages.append(f"S2 매핑 입력 차단 대상입니다: {result.spec.s2_gate}")
    if result.rows.empty:
        messages.append("어댑터가 데이터 행을 만들지 못했습니다.")
    elif result.default_feed_rows.empty:
        messages.append("파싱은 됐지만 S2 매핑으로 보낼 입력 행이 없습니다.")
    if any(audit.status == "header_not_found" for audit in result.sheet_audits):
        failed = ", ".join(audit.sheet for audit in result.sheet_audits if audit.status == "header_not_found")
        messages.append(f"헤더를 찾지 못한 시트가 있습니다: {failed}. 시트명, 헤더 위치, 헤더명이 바뀌었는지 확인하세요.")
    return messages


def adapter_warning_messages(result: NormalizationResult) -> list[str]:
    messages: list[str] = []
    if result.default_feed_rows.empty:
        return messages

    parsed_rows = len(result.rows)
    feed_rows = len(result.default_feed_rows)
    if parsed_rows != feed_rows:
        messages.append(f"파싱 행 {parsed_rows:,}개 중 S2 매핑 입력은 {feed_rows:,}개입니다. 제외/검토 규칙을 확인하세요.")
    if not result.spec.s2_amount_policy_locked:
        messages.append(f"S2 금액 4컬럼 출력은 아직 잠금 전입니다: {result.spec.s2_gate}")
    if result.default_feed_rows[STANDARD_TITLE_COLUMN].map(text).eq("").any():
        messages.append("상품명 빈 행이 남아 있습니다. 제목 컬럼 감지 또는 행 필터를 확인하세요.")
    return messages


def list_platforms() -> list[str]:
    return sorted(REGISTRY)


def detect_platform(source_name: str | Path) -> str | None:
    haystack = _norm(str(source_name))
    for platform, aliases in _PLATFORM_ALIASES.items():
        for alias in aliases:
            if _norm(alias) and _norm(alias) in haystack:
                return platform
    for platform in sorted(REGISTRY, key=len, reverse=True):
        if _norm(platform) in haystack:
            return platform
    return None


def normalize_settlement(
    source: str | Path | BinaryIO | io.BytesIO,
    *,
    platform: str | None = None,
    source_name: str | None = None,
) -> NormalizationResult:
    display_name = source_name or getattr(source, "name", None) or str(source)
    selected_platform = platform or detect_platform(display_name)
    if not selected_platform:
        raise ValueError(f"플랫폼을 감지하지 못했습니다. 플랫폼을 직접 선택해 주세요: {display_name}")
    if selected_platform not in REGISTRY:
        raise ValueError(f"등록되지 않은 플랫폼입니다: {selected_platform}")

    spec = REGISTRY[selected_platform]
    file_status = _file_status(spec, display_name)
    if file_status == "excluded_by_rule":
        return NormalizationResult(
            platform=selected_platform,
            source_name=display_name,
            file_status=file_status,
            spec=spec,
            rows=_empty_rows(),
            sheet_audits=[SheetAudit(sheet="", status="excluded_by_rule", note=spec.exclude_rule)],
        )

    workbook = _load_workbook(source)
    parsed: list[pd.DataFrame] = []
    audits: list[SheetAudit] = []
    for sheet in workbook.worksheets:
        if not _sheet_in_scope(spec, sheet.title):
            audits.append(SheetAudit(sheet=sheet.title, status="excluded_sheet", note=spec.exclude_rule))
            continue
        frame, audit = _parse_sheet(sheet, spec, display_name, file_status)
        audits.append(audit)
        if not frame.empty:
            parsed.append(frame)

    rows = pd.concat(parsed, ignore_index=True) if parsed else _empty_rows()
    return NormalizationResult(
        platform=selected_platform,
        source_name=display_name,
        file_status=file_status,
        spec=spec,
        rows=rows,
        sheet_audits=audits,
    )


def summarize_normalization(result: NormalizationResult) -> dict[str, Any]:
    rows = result.rows
    title_present = int(rows[STANDARD_TITLE_COLUMN].map(text).ne("").sum()) if STANDARD_TITLE_COLUMN in rows else 0
    return {
        "platform": result.platform,
        "source_name": result.source_name,
        "file_status": result.file_status,
        "parser_contract": result.spec.parser_contract,
        "parsed_rows": len(rows),
        "title_present_rows": title_present,
        "default_feed_rows": len(result.default_feed_rows),
        "amount_rule_status": result.spec.amount_rule_status,
        "s2_amount_policy_locked": result.spec.s2_amount_policy_locked,
        "s2_gate": result.spec.s2_gate,
    }


def _load_workbook(source: str | Path | BinaryIO | io.BytesIO):
    if hasattr(source, "seek"):
        source.seek(0)
    try:
        return load_workbook(source, data_only=True, read_only=False)
    except Exception:
        if hasattr(source, "seek"):
            source.seek(0)
        return _load_workbook_values_only_ooxml(source)


def _load_workbook_values_only_ooxml(source: str | Path | BinaryIO | io.BytesIO) -> _FallbackWorkbook:
    with _zipfile_from_source(source) as archive:
        shared_strings = _read_shared_strings(archive)
        workbook_root = _read_xml_from_archive(archive, "xl/workbook.xml")
        rels = _read_workbook_relationships(archive)
        worksheets: list[_FallbackWorksheet] = []
        for sheet_node in workbook_root.findall("x:sheets/x:sheet", OOXML_NS):
            title = sheet_node.attrib.get("name", "")
            rel_id = sheet_node.attrib.get(f"{{{OOXML_REL_NS}}}id", "")
            target = rels.get(rel_id, "")
            if not target:
                continue
            sheet_path = _normalize_ooxml_path("xl", target)
            if sheet_path not in archive.namelist():
                continue
            sheet_root = _read_xml_from_archive(archive, sheet_path)
            values, merged_ranges = _read_sheet_values(sheet_root, shared_strings)
            worksheets.append(
                _FallbackWorksheet(
                    title=title,
                    values=values,
                    merged_cells=_FallbackMergedCells(merged_ranges),
                )
            )
    if not worksheets:
        raise ValueError("OOXML fallback failed: no worksheets could be read")
    return _FallbackWorkbook(worksheets=worksheets)


def _zipfile_from_source(source: str | Path | BinaryIO | io.BytesIO) -> zipfile.ZipFile:
    if isinstance(source, (str, Path)):
        return zipfile.ZipFile(source)
    if hasattr(source, "seek"):
        source.seek(0)
    payload = source.read()
    if isinstance(payload, str):
        payload = payload.encode()
    return zipfile.ZipFile(io.BytesIO(payload))


def _read_xml_from_archive(archive: zipfile.ZipFile, name: str) -> ET.Element:
    try:
        return ET.fromstring(archive.read(name))
    except KeyError as exc:
        raise ValueError(f"OOXML fallback failed: missing {name}") from exc
    except ET.ParseError as exc:
        raise ValueError(f"OOXML fallback failed: invalid XML {name}: {exc}") from exc


def _read_workbook_relationships(archive: zipfile.ZipFile) -> dict[str, str]:
    root = _read_xml_from_archive(archive, "xl/_rels/workbook.xml.rels")
    result: dict[str, str] = {}
    for rel in root.findall("pr:Relationship", OOXML_NS):
        rel_id = rel.attrib.get("Id", "")
        target = rel.attrib.get("Target", "")
        if rel_id and target:
            result[rel_id] = target
    return result


def _normalize_ooxml_path(base_dir: str, target: str) -> str:
    target = target.replace("\\", "/")
    if target.startswith("/"):
        return target.lstrip("/")
    parts: list[str] = []
    for part in f"{base_dir}/{target}".split("/"):
        if part in {"", "."}:
            continue
        if part == "..":
            if parts:
                parts.pop()
            continue
        parts.append(part)
    return "/".join(parts)


def _read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = _read_xml_from_archive(archive, "xl/sharedStrings.xml")
    result: list[str] = []
    for item in root.findall("x:si", OOXML_NS):
        result.append("".join(node.text or "" for node in item.findall(".//x:t", OOXML_NS)))
    return result


def _read_sheet_values(root: ET.Element, shared_strings: list[str]) -> tuple[list[list[Any]], list[_FallbackMergedRange]]:
    cells: dict[tuple[int, int], Any] = {}
    max_row = 0
    max_col = 0
    for row_node in root.findall(".//x:sheetData/x:row", OOXML_NS):
        row_idx = int(row_node.attrib.get("r", "0") or 0)
        if row_idx <= 0:
            row_idx = max_row + 1
        sequential_col = 0
        for cell_node in row_node.findall("x:c", OOXML_NS):
            ref = cell_node.attrib.get("r", "")
            if ref:
                parsed_row, col_idx = _split_cell_ref(ref)
                row_idx = parsed_row or row_idx
            else:
                sequential_col += 1
                col_idx = sequential_col
            value = _read_cell_value(cell_node, shared_strings)
            cells[(row_idx, col_idx)] = value
            max_row = max(max_row, row_idx)
            max_col = max(max_col, col_idx)

    values = [[None for _ in range(max_col)] for _ in range(max_row)]
    for (row_idx, col_idx), value in cells.items():
        if row_idx > 0 and col_idx > 0:
            values[row_idx - 1][col_idx - 1] = value

    merged_ranges = []
    for merge_node in root.findall(".//x:mergeCells/x:mergeCell", OOXML_NS):
        ref = merge_node.attrib.get("ref", "")
        bounds = _range_bounds(ref)
        if bounds is not None:
            merged_ranges.append(_FallbackMergedRange(bounds=bounds))
    return values, merged_ranges


def _read_cell_value(cell_node: ET.Element, shared_strings: list[str]) -> Any:
    cell_type = cell_node.attrib.get("t", "")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell_node.findall(".//x:t", OOXML_NS))
    value_node = cell_node.find("x:v", OOXML_NS)
    raw = value_node.text if value_node is not None else ""
    if cell_type == "s":
        try:
            return shared_strings[int(raw)]
        except (ValueError, IndexError):
            return ""
    if cell_type == "str":
        return raw or ""
    if cell_type == "b":
        return raw == "1"
    if raw in (None, ""):
        return ""
    try:
        number = float(raw)
        return int(number) if number.is_integer() else number
    except ValueError:
        return raw


def _split_cell_ref(ref: str) -> tuple[int, int]:
    match = re.fullmatch(r"([A-Za-z]+)(\d+)", ref)
    if not match:
        return 0, 0
    return int(match.group(2)), _column_letters_to_number(match.group(1))


def _range_bounds(ref: str) -> tuple[int, int, int, int] | None:
    if ":" not in ref:
        return None
    start, end = ref.split(":", 1)
    min_row, min_col = _split_cell_ref(start)
    max_row, max_col = _split_cell_ref(end)
    if not all([min_row, min_col, max_row, max_col]):
        return None
    return min_col, min_row, max_col, max_row


def _column_letters_to_number(value: str) -> int:
    number = 0
    for char in value.upper():
        if not ("A" <= char <= "Z"):
            return 0
        number = number * 26 + (ord(char) - ord("A") + 1)
    return number


def _empty_rows() -> pd.DataFrame:
    return pd.DataFrame(columns=STANDARD_COLUMNS)


def _file_status(spec: AdapterSpec, source_name: str) -> str:
    normalized = _norm(source_name)
    if spec.blocks_default_feed:
        return "blocked_non_s2_source"
    if spec.platform == "부커스" and "복사본" in source_name:
        return "excluded_by_rule"
    if spec.platform == "네이버" and "통합" in source_name:
        return "review_gate_not_default"
    if "확장체크" in source_name:
        return "review_gate_not_default"
    return "include"


def _sheet_in_scope(spec: AdapterSpec, sheet_name: str) -> bool:
    name = _norm(sheet_name)
    if spec.platform == "윌라" and "콘텐츠가격변동이력" in name:
        return False
    if spec.platform == "판무림":
        return "세부내역" in name
    if spec.platform == "스토린랩":
        return name in {_norm("원스"), _norm("원스(북패스)"), _norm("무툰")}
    if spec.platform == "북팔":
        return name != _norm("Sheet")
    if spec.platform == "미스터블루":
        return "작품별" in name
    if spec.platform == "보인&국립장애인도서관":
        return "1차구매" in name
    if spec.platform in {"알라딘 종이책", "한아름"}:
        return True

    rule = spec.source_sheet_rule.strip()
    if not rule or rule in {"없음", "날짜범위 시트", "월별 정산내역 시트", "상세 판매일 시트"}:
        return True
    tokens = [tok.strip() for tok in re.split(r"[,，]| 및 | 또는 ", rule) if tok.strip()]
    for token in tokens:
        token = token.replace("시트", "").replace("우선", "").strip()
        if not token:
            continue
        if "*" in token and fnmatch.fnmatch(sheet_name, token):
            return True
        if _norm(token) in name:
            return True
    return False


def _parse_sheet(sheet: Worksheet, spec: AdapterSpec, source_name: str, file_status: str) -> tuple[pd.DataFrame, SheetAudit]:
    rows = _sheet_rows(sheet)
    if spec.parser_contract == "multi_section_repeated_header_parser":
        frame = _parse_repeated_sections(rows, spec, source_name, sheet.title, file_status)
        return frame, SheetAudit(
            sheet=sheet.title,
            status="parsed" if not frame.empty else "no_data",
            header_row="repeated",
            data_start="section+1",
            parsed_rows=len(frame),
            title_present_rows=_title_count(frame),
        )

    header_idx = _find_header_row(rows, spec)
    if header_idx is None:
        return _empty_rows(), SheetAudit(sheet=sheet.title, status="header_not_found", note=spec.parser_contract)

    headers = _unique_headers(_header_values(rows, header_idx))
    data_rows = rows[header_idx + 1 :]
    raw = pd.DataFrame(data_rows, columns=headers)
    frame = _standardize(raw, spec, source_name, sheet.title, header_idx, file_status)
    return frame, SheetAudit(
        sheet=sheet.title,
        status="parsed",
        header_row=str(header_idx + 1),
        data_start=str(header_idx + 2),
        parsed_rows=len(frame),
        title_present_rows=_title_count(frame),
    )


def _parse_repeated_sections(
    rows: list[list[Any]],
    spec: AdapterSpec,
    source_name: str,
    sheet_name: str,
    file_status: str,
) -> pd.DataFrame:
    header_indexes = []
    for idx, row in enumerate(rows):
        if _header_score(row, spec) >= 8:
            header_indexes.append(idx)
    frames: list[pd.DataFrame] = []
    for seq, header_idx in enumerate(header_indexes):
        next_idx = header_indexes[seq + 1] if seq + 1 < len(header_indexes) else len(rows)
        headers = _unique_headers(_header_values(rows, header_idx))
        raw = pd.DataFrame(rows[header_idx + 1 : next_idx], columns=headers)
        frame = _standardize(raw, spec, source_name, sheet_name, header_idx, file_status)
        if not frame.empty:
            frames.append(frame)
    return pd.concat(frames, ignore_index=True) if frames else _empty_rows()


def _sheet_rows(sheet: Worksheet) -> list[list[Any]]:
    values = [list(row) for row in sheet.iter_rows(values_only=True)]
    if not values:
        return []
    for merged in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        top_left = values[min_row - 1][min_col - 1]
        for row_idx in range(min_row - 1, max_row):
            for col_idx in range(min_col - 1, max_col):
                if values[row_idx][col_idx] in (None, ""):
                    values[row_idx][col_idx] = top_left
    return values


def _find_header_row(rows: list[list[Any]], spec: AdapterSpec) -> int | None:
    best_idx = None
    best_score = -1
    limit = min(len(rows), 100)
    for idx in range(limit):
        score = _header_score(rows[idx], spec)
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx if best_score >= 8 else None


def _header_score(row: list[Any], spec: AdapterSpec) -> int:
    normalized_cells = {_norm(str(cell)) for cell in row if text(cell)}
    title_hit = any(_norm(candidate) in normalized_cells for candidate in spec.title_candidates)
    if not title_hit:
        return 0
    score = 8
    for candidates, weight in [
        (spec.author_candidates, 2),
        (spec.external_id_candidates, 1),
        (spec.sale_amount_candidates, 2),
        (spec.settlement_amount_candidates, 2),
        (spec.offset_amount_candidates, 1),
    ]:
        if any(_norm(candidate) in normalized_cells for candidate in candidates):
            score += weight
    return score


def _header_values(rows: list[list[Any]], header_idx: int) -> list[str]:
    return [text(value) for value in rows[header_idx]]


def _unique_headers(headers: list[str]) -> list[str]:
    result: list[str] = []
    seen: dict[str, int] = {}
    for idx, header in enumerate(headers, start=1):
        base = header or f"column_{idx}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        result.append(base if count == 1 else f"{base}__{count}")
    return result


def _standardize(
    raw: pd.DataFrame,
    spec: AdapterSpec,
    source_name: str,
    sheet_name: str,
    header_idx: int,
    file_status: str,
) -> pd.DataFrame:
    if raw.empty:
        return _empty_rows()

    raw = _drop_platform_blocked_raw_columns(raw, spec)

    title = _pick_series(raw, spec.title_candidates)
    title_text = title.map(text)
    data_mask = title_text.ne("") & ~title_text.map(_is_non_data_title)
    data = raw[data_mask].copy()
    if data.empty:
        return _empty_rows()

    title = _pick_series(data, spec.title_candidates)
    out = pd.DataFrame(index=data.index)
    out["platform"] = spec.platform
    out["source_file"] = source_name
    out["source_sheet"] = sheet_name
    out["source_row"] = data.index.map(lambda idx: int(idx) + header_idx + 2)
    out["file_status"] = file_status
    out["row_status"] = "data"
    out[STANDARD_TITLE_COLUMN] = title.map(text)
    out["작가명"] = _pick_series(data, spec.author_candidates).map(text)
    out["외부콘텐츠ID"] = _pick_series(data, spec.external_id_candidates).map(text)
    out["판매금액_후보"] = _pick_series(data, spec.sale_amount_candidates).map(_number_or_blank)
    out["정산기준액_후보"] = _pick_series(data, spec.settlement_amount_candidates).map(_number_or_blank)
    out["상계금액_후보"] = _pick_series(data, spec.offset_amount_candidates).map(_number_or_blank)
    out["정제_상품명"] = out[STANDARD_TITLE_COLUMN].map(clean_title)
    out["parser_contract"] = spec.parser_contract
    out["amount_rule_status"] = spec.amount_rule_status
    out["s2_gate"] = spec.s2_gate
    return out.reset_index(drop=True)


def _drop_platform_blocked_raw_columns(raw: pd.DataFrame, spec: AdapterSpec) -> pd.DataFrame:
    keywords = _PLATFORM_BLOCKED_RAW_COLUMN_KEYWORDS.get(spec.platform, ())
    if not keywords:
        return raw

    blocked = []
    for column in raw.columns:
        base = re.sub(r"__\d+$", "", str(column))
        normalized = _norm(base)
        if any(keyword in normalized for keyword in keywords):
            blocked.append(column)
    if not blocked:
        return raw
    return raw.drop(columns=blocked)


def _pick_series(df: pd.DataFrame, candidates: Iterable[str]) -> pd.Series:
    candidate_norms = [_norm(candidate) for candidate in candidates if _norm(candidate)]
    for column in df.columns:
        column_norm = _norm(re.sub(r"__\d+$", "", str(column)))
        if column_norm in candidate_norms:
            return df[column]
    return pd.Series([""] * len(df), index=df.index, dtype=object)


def _title_count(frame: pd.DataFrame) -> int:
    if frame.empty or STANDARD_TITLE_COLUMN not in frame:
        return 0
    return int(frame[STANDARD_TITLE_COLUMN].map(text).ne("").sum())


def _is_non_data_title(value: str) -> bool:
    normalized = _norm(value)
    return normalized in {
        "",
        "작품명",
        "상품명",
        "채널상품명",
        "콘텐츠제목",
        "도서명",
        "제목",
        "타이틀",
        "콘텐츠명",
        "컨텐츠",
        "합계",
        "총합계",
        "총액",
        "총계",
        "소계",
        "sum",
        "total",
    } or any(marker in normalized for marker in ["소계", "총합", "원천징수"])


def _number_or_blank(value: Any) -> Any:
    value_text = text(value)
    if not value_text:
        return ""
    cleaned = value_text.replace(",", "").replace("원", "").replace(" ", "")
    if re.fullmatch(r"-?\d+(?:\.\d+)?", cleaned):
        number = float(cleaned)
        return int(number) if number.is_integer() else number
    return value


def _field_candidates(field_text: str) -> list[str]:
    if not field_text or field_text in {"없음", "없음 확인 필요", "정산기준액 없음"}:
        return []
    if "거래내용" in field_text:
        return ["거래내용"]
    text_value = field_text
    for token in ["후보", "정책 필요", "필요", "있는 그룹만", "직접", "그 외", "상단", "기반", "적용 여부"]:
        text_value = text_value.replace(token, "")
    text_value = re.sub(r"\s+x\s+", "/", text_value)
    text_value = re.sub(r"\s*\+\s*", "/", text_value)
    text_value = text_value.replace(" 또는 ", "/").replace(" 및 ", "/").replace(",", "/")
    parts = [part.strip(" .") for part in text_value.split("/") if part.strip(" .")]
    result: list[str] = []
    for part in parts:
        part = re.sub(r"\s*\([^)]*불명[^)]*\)", "", part).strip()
        part = re.sub(r"\s*컬럼 선택 기준 확정 후.*$", "", part).strip()
        if not part or any(bad in part for bad in ["정책", "확정", "없음", "아님"]):
            continue
        result.append(part)
    return result


def _norm(value: str) -> str:
    value = unicodedata.normalize("NFKC", value or "").lower()
    return re.sub(r"[\s_\-–—/\\:;.,()[\]{}<>〈〉《》'\"`|+]+", "", value)


_PLATFORM_ALIASES = {
    "교보": ["교보", "교보문고"],
    "리디북스": ["리디북스", "리디"],
    "무툰": ["무툰"],
    "미스터블루": ["미스터블루"],
    "밀리의서재": ["밀리의서재"],
    "보인&국립장애인도서관": ["보인&국립장애인도서관", "국립장애인도서관", "보인"],
    "토스(구루컴퍼니)": ["토스", "구루컴퍼니"],
    "피우리(누온)": ["피우리", "누온"],
}


_REGISTRY_ROWS = [
    ("교보", "single_header_with_leading_merged_title", "single_header_policy_gate", "Sheet1", "", "상품명", "저자", "판매상품ID / epub isbn / pdf isbn / 종이책ISBN / 북넘버", "정산대상판매가총액 또는 판매가 x 판매수량", "정산액", "없음 확인 필요", "needs_reconcile", "대표월 fixture + 판매금액 산식 확정 전 S2 출력 금지"),
    ("구글", "google_transaction_fixed_columns", "special_parser_required", "GoogleSalesTransactionReport", "", "Title", "Author", "Id / Product / Primary ISBN", "Publisher Revenue / Payment Amount 후보", "법인세 차감 후 금액 또는 Publisher Revenue 후보", "미국 원천징수세 / 법인세 차감액 후보", "needs_policy", "Google 금액 기준 확정 + 외화/세금 처리 전 S2 출력 금지"),
    ("네이버", "two_row_merged_header_flatten", "special_parser_required", "contentsSelling_*", "복사본, 중복 통합본은 review gate", "컨텐츠", "작가명", "컨텐츠No / 공급자코드", "합계", "정산금액 있는 그룹만 직접 후보, 그 외 정책 필요", "마켓수수료(추정치) / 유상 이용권 보정", "needs_policy", "2행 헤더 flatten + 통합/선투자 중복 검산 전 S2 출력 금지"),
    ("노벨피아", "single_header", "single_header_policy_gate", "일별 정산", "", "상품명", "작가명", "작품코드", "판매합계 또는 판매금액", "정산금액", "취소금액", "needs_cancel_policy", "취소금액 처리 정책 확정 후 S2 출력"),
    ("로망띠끄", "single_header_row5_with_merged_banner", "single_header_policy_gate", "styleB(바로북)*", "", "도서명", "저자", "도서코드", "판매액", "정산액", "없음 확인 필요", "candidate_confirmed_after_reconcile", "대표월 fixture + 총액 reconcile 후 출력"),
    ("리디북스", "wide_single_header_limited_columns", "wide_header_policy_gate", "calculate_1 및 리디 정산상세 시트", "", "제목 / 시리즈명", "저자", "도서 ID / 시리즈 ID / 전자책ISBN10/13", "판매액 + 단권/세트/대여 판매액 후보", "정산액 / 앱마켓 정산대상액 후보", "취소액 / 앱마켓 수수료 / 앱마켓 취소액", "needs_policy", "판매/취소/앱마켓 산식 확정 전 S2 출력 금지"),
    ("모픽", "single_header_variants", "single_header_policy_gate", "작품별정산", "", "작품명", "작가명", "없음", "총 매출액 또는 순 매출액", "정산액", "총 매출액-순 매출액 후보", "needs_policy", "총/순 매출 기준 확정 후 S2 출력"),
    ("무툰", "merged_header_coin_table", "single_header_amount_policy_required", "Sheet", "", "타이틀", "작가", "없음", "합계 / 사용코인 후보", "정산총액 또는 정산금액", "공제수수료 / 취소코인", "needs_coin_policy", "코인-원화 및 취소/공제 기준 확정 전 S2 출력 금지"),
    ("문피아", "single_header_limited_columns", "single_header_policy_gate", "다우인큐브", "", "작품", "작가", "작품코드", "총매출", "정산", "구매취소 / 대여취소", "needs_cancel_policy", "취소/IOS/Google 포함 여부 확정 후 S2 출력"),
    ("미소설", "single_header_variants", "single_header_policy_gate", "cpexcel*", "확장체크 파일은 fixture 제외", "타이틀", "작가명", "작품번호", "전체매출(원)", "총지급액(원) 또는 지급액(원) ASP", "결제수수료(원)", "needs_policy", "지급액 컬럼 선택 기준 확정 후 S2 출력"),
    ("미스터블루", "sheet_whitelist_workbook", "mixed_sheet_policy_gate", "작품별 우선, 볼륨별은 검산/보조 후보", "정산기준 없는 보조 시트는 출력 제외", "작품명", "작가명", "작품코드", "작품별 합계(정액+종량) / 볼륨별 소계 후보", "정산기준액 후보 없음. 정책 필요", "없음 확인 필요", "blocked_until_settlement_basis", "정산기준액 정의 전 S2 출력 금지"),
    ("밀리의서재", "single_header_variants", "single_header_policy_gate", "list", "", "콘텐츠명", "저자명", "전자출판물 ISBN / 유통사 상품코드", "발생 금액", "정산 예정 금액 / 정산 금액", "없음 확인 필요", "candidate_confirmed_after_reconcile", "에피소드명 집계 단위 확정 + 총액 reconcile 후 출력"),
    ("보인&국립장애인도서관", "purchase_selection_list", "non_s2_source_blocked", "없음", "목록선정*", "제목", "저자명", "없음", "판매가 / 구매비", "정산기준액 없음", "없음", "blocked_non_sales", "S2 판매상세가 아니라 구매/목록선정 자료로 판단. 기본 출력 금지"),
    ("부커스", "single_header_with_duplicate_file_filter", "single_header_amount_policy_required", "CP 콘텐츠별 정산", "- 복사본 파일 제외", "콘텐츠 제목", "저자명", "ISBN", "정가 x 열람 횟수 후보", "정산 금액(원)", "정산율 기반 파생 후보", "needs_derived_sale_policy", "판매금액 파생식 확정 전 S2 출력 금지"),
    ("북큐브", "single_header_row2", "single_header_policy_gate", "Sheet1", "", "제목", "저자", "도서번호 / ISBN / e-ISBN(PDF) / e-ISBN(ePub) / 제휴도서번호", "판매액", "정산액 / 정산대상금액", "할인 / 수수료", "candidate_confirmed_after_reconcile", "대표월 fixture + 할인/수수료 상계 기준 확정 후 출력"),
    ("북팔", "sheet_whitelist_plus_legacy_exclusion", "mixed_sheet_policy_gate", "마켓, 날짜범위 시트", "Sheet 안내문/전자계산서 본문 시트 제외", "작품명", "필명 / 닉네임", "아이디 / 순서 후보", "매출", "수익금 / 수익", "없음 확인 필요", "candidate_confirmed_after_reconcile", "보조 안내문 시트 제외 + 총액 reconcile 후 출력"),
    ("블라이스", "single_header_with_merged_cells", "single_header_policy_gate", "CP정산*", "", "작품명", "작가명", "작품NO", "판매액", "정산금액", "수수료", "candidate_confirmed_after_reconcile", "수수료 컬럼명 정규화 + 총액 reconcile 후 출력"),
    ("스낵북", "single_header", "single_header_policy_gate", "settle_list_*", "", "작품명", "필자명", "작품ID", "판매(원)", "정산(원)", "없음 확인 필요", "candidate_confirmed_after_reconcile", "대표월 fixture + 총액 reconcile 후 출력"),
    ("스토린랩", "schema_signature_subadapters", "multi_adapter_policy_gate", "원스, 원스(북패스), 무툰 시트", "정리 시트(B2BC 집계형) 제외 또는 별도 수동", "채널상품명 / 상품명 / 이용상품명 / 타이틀", "글작가 / 작가", "채널상품ID / 상품ID / 키다리코드 / 이용상품ID", "판매금액 / 정액권 총매출액 / 판매총액 / 합계금액", "정산액 / 입금액 / 상품별 정산금액", "취소금액 / 차감 / 취소코인", "needs_subadapter_policy", "서브어댑터별 amount rule 확정 전 S2 출력 금지"),
    ("신영미디어", "single_header", "single_header_policy_gate", "날짜범위 시트", "", "제목", "저자", "번호", "합계", "소득액", "없음 확인 필요", "candidate_confirmed_after_reconcile", "대표월 fixture + 총액 reconcile 후 출력"),
    ("알라딘", "single_header", "single_header_policy_gate", "sales_*", "", "제목", "저자명", "ItemId / ISBN / CID", "판매가", "정산액", "판매형태=취소/원주문일시 기반 취소 처리", "needs_cancel_policy", "판매/취소 행 처리 정책 확정 후 S2 출력"),
    ("알라딘 종이책", "ledger_not_content_sales", "non_s2_source_blocked", "없음", "거래처별 거래원장", "거래내용 후보이나 작품명 아님", "없음", "없음", "출고금액 / 반품금액", "현잔액 등 원장 금액", "반품금액", "blocked_non_content_ledger", "전자 S2 콘텐츠 판매상세 아님. 기본 출력 금지"),
    ("에이블리", "single_header_plus_summary_exclusion", "mixed_sheet_policy_gate", "상세 판매일 시트", "행 레이블 pivot/요약 시트는 출력 제외", "작품명", "작가명", "작품 ID", "판매 금액 합계 (원)", "정산 금액 합계 (원)", "운영 수수료 합계 (원)", "candidate_confirmed_after_reconcile", "요약 시트 제외 + 총액 reconcile 후 출력"),
    ("에피루스", "single_header", "single_header_policy_gate", "월별 정산내역 시트", "", "제목", "저자", "없음", "판매금액", "정산액", "수수료", "candidate_confirmed_after_reconcile", "대표월 fixture + 수수료 상계 정책 확인 후 출력"),
    ("예스24", "single_header", "single_header_policy_gate", "Sheet1", "", "도서명", "저자명", "bookID / ePubID / 전자책ISBN / 종이책ISBN / 세트코드", "서점판매가 또는 출판사판매가", "출판사정산액", "서점환불가 / 환불일 기반 취소 처리", "needs_policy", "판매가 기준/환불 처리 정책 확정 후 출력"),
    ("올툰", "single_header", "single_header_amount_policy_required", "시트1", "", "작품명", "없음", "없음", "총 매출액(원) / 코인 사용수량", "정산 대상 금액(수수료 제외)", "앱스토어 수수료(원) / 올웨이즈 수수료(원)", "needs_fee_policy", "수수료 제외/순매출 기준 확정 후 S2 출력"),
    ("원스토어", "wide_single_header_limited_columns", "wide_header_policy_gate", "multimedia", "", "채널상품명 / 상품명", "글작가", "채널상품ID / 상품ID / 파트너ID", "판매 / 합계 / 정액제 정산대상액 후보", "정산지급액", "취소 / 앱마켓수수료 / 서비스이용료", "needs_policy", "판매/정액제/취소/수수료 산식 확정 전 S2 출력 금지"),
    ("윌라", "sheet_whitelist_single_header", "single_header_with_aux_exclusion", "sheet", "콘텐츠 가격 변동 이력 시트 제외", "콘텐츠명", "저자", "코드 / 전자책ISBN", "공급가 또는 정가 후보", "정산 금액", "없음 확인 필요", "needs_policy", "공급가/정가 판매금액 기준 확정 + 가격변동이력 제외 후 출력"),
    ("조아라", "schema_signature_variants", "multi_schema_amount_policy_required", "작품별 정산리스트*, 후원쿠폰*", "", "작품명", "작가명", "작품코드 있는 그룹만", "단가 x 판매건수 / 금액 / 이용권 후보. 일부 단행본은 판매금액 불명", "정산금액", "없음 확인 필요", "needs_derived_sale_policy", "단가 없는 그룹의 판매금액 정책 확정 전 S2 출력 금지"),
    ("카카오", "merged_header_single_table", "wide_header_policy_gate", "정산리포트_카카오페이지_*", "미발행/선투자 파일은 중복 검토", "시리즈명", "작가명", "시리즈ID / 제품코드 / 계약UID", "공급대가 또는 총합계-순매출 후보", "공급가액", "세액은 별도 보존. 상계 아님", "needs_tax_policy", "과/면세 및 공급가액/공급대가 기준 확정 전 출력 금지"),
    ("큐툰", "merged_header_coin_table", "single_header_amount_policy_required", "Sheet", "", "타이틀", "작가", "없음", "합계 / 사용코인 후보", "정산총액", "공제수수료 / 취소코인", "needs_coin_policy", "코인-원화 및 취소/공제 기준 확정 전 S2 출력 금지"),
    ("토스(구루컴퍼니)", "single_header_variants", "single_header_policy_gate", "정산_공급사_*", "", "작품명", "작가", "작품번호", "콘텐츠 매출금액", "콘텐츠 정산금액 또는 면세+과세 합산", "결제수수료", "needs_tax_split_policy", "면세/과세 합산 기준 확정 후 출력"),
    ("판무림", "cover_exclusion_contextual_amount_flatten", "special_parser_required", "세부내역", "표지 시트 제외", "작품 제목 / 회차 제목", "저자", "시리즈 코드 / 각 권 코드", "소장/대여/정액제/포인트 구간 판매금액 flatten 후 합산 후보", "표지 정산비율 적용 여부 정책 필요", "포인트 사용 후보", "needs_contextual_amount_policy", "상단 문맥+반복 판매금액 flatten 전 S2 출력 금지"),
    ("피우리(누온)", "single_header", "single_header_policy_gate", "Worksheet", "", "제목", "작가", "피우리상품번호 / CP관리번호", "매출액 또는 판매가 x 판매수", "정산액", "없음 확인 필요", "candidate_confirmed_after_reconcile", "대표월 fixture + 판매가/매출액 기준 확인 후 출력"),
    ("피플앤스토리", "single_header_with_small_sample_gate", "single_header_policy_gate", "다운로드", "", "작품명", "작가명", "ISBN / 관리코드", "판매금액(원)", "정산금액(원)", "없음 확인 필요", "candidate_confirmed_after_reconcile", "샘플 행 적음. 다른 월 1개 추가 fixture 후 출력"),
    ("하이북", "single_header", "single_header_policy_gate", "정산리스트", "", "작품명", "작가명", "prodSq", "판매금액(구매)+판매금액(대여) 또는 판매금액", "정산금액(구매)+정산금액(대여) 또는 정산금액", "없음 확인 필요", "needs_purchase_rental_policy", "구매/대여 합산 단위 확정 후 출력"),
    ("한아름", "multi_section_repeated_header_parser", "special_parser_required", "건당 로그/시간 로그 반복 섹션", "요약, 계산서 안내, 합계, 반복 헤더 row 제외", "작품명", "작가명", "BOOK NO", "건당 로그 금액 / 시간 로그 히트 기반 정책 필요", "상단 정산금액 요약 및 배분률 대조", "배분률 기반 차액 후보", "needs_section_policy", "구간 파서 + 시간/건당 과금 정책 확정 전 S2 출력 금지"),
]


REGISTRY: dict[str, AdapterSpec] = {
    row[0]: AdapterSpec(*row)
    for row in _REGISTRY_ROWS
}
